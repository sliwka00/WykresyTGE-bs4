import matplotlib.pyplot as plt
import pandas as pd
import datetime as dt
import time
import numpy as np
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import holidays
import holidays.countries
import streamlit as st
from bs4 import BeautifulSoup
import requests
import plotly.graph_objects as go
from plotly.subplots import make_subplots

cale=[]
kwartaly=[]
msc=[]
print(cale)
df = pd.read_excel(r'abc.xlsx')


df = df.replace('-',np.nan)   #zamienia  "-" na Nan w komórkach gdzie nie ma ceny
df = df.astype({'DKR':float})  #zamienia kolumne DKR na floaty (dane były jako string)
df['kontrakt short'] = df['Kontrakt'].str.split("_").str[-1]       #Skraca nazwe kontraktu do uniwersalnego (dla base i peak) żeby je sparować
df['Data']=pd.to_datetime(df['Data'], format='%d-%m-%Y')
df['wolumen'] = [float(str(val).replace(u'\xa0','').replace(',','.')) for val in df['wolumen'].values]   #wyrzucenie dziwnych znaków z wolumenu i zamiana na float
df3 = df[['Data','DKR','typ','wolumen','kontrakt short']]  #stworzenie skróconego df bez zbędnych kolumn
df_base = df3[df3['typ'] == 'BASE']     #stworzenie df dla base
df_peak = df3[df3['typ'] == 'PEAK']
df_wsp = pd.merge(df_base,df_peak, on=['Data','kontrakt short'])  #połączenie df_base i df_peak dzieki temu można dodać kolumne ratio
df_wsp['ratio']=df_wsp['DKR_y']/df_wsp['DKR_x']  #kolumna z ratio

# Pętla do uzupełniania listy produktów, które znajdują sie w pliku zródłowym
for produkt in df['Kontrakt']:
    if "_Y-" in produkt:
        if produkt not in cale:
            cale.append(produkt)
    elif "_Q-" in produkt:
        if produkt not in kwartaly:
            kwartaly.append(produkt)
    elif "_M-" in produkt:
        if produkt not in msc:
            msc.append(produkt)
    else:
        continue

msc.sort()
kwartaly.sort()
cale.sort()
print(cale)
print(kwartaly)
print(msc)

st.set_option('deprecation.showPyplotGlobalUse', False)  #usuniecie komunikatu ze strony

# Funkcja rysujaca wykres w matplotlib
def draw_chart(produkt):
    df2=df[df['Kontrakt']==produkt]
    data=df2['Data']
    cena=(df2['DKR'])
    wolumen=(df2['liczba transakcji'])
    # Tworzenie figury i osi
    fig, ax1 = plt.subplots()
    ax1.bar(data, wolumen, color='gray', alpha=0.5)
    ax1.set_ylabel('liczba transakcji')
    ax2 = ax1.twinx()

    ax2.plot(data, cena, marker='o', linestyle='-', color='blue',markersize=2, markerfacecolor='black')
    ax2.set_title(produkt)
    ax2.set_xlabel('Data')
    ax2.set_ylabel('cena')
    #
    #myFmt = mdates.DateFormatter('%d-%m-%Y')
    fig.autofmt_xdate(rotation=35, ha='right')    #rotuje daty wyświetlane pod wykresem

    st.pyplot()

def draw_interactive(produkt):    # na próbe z plotly, ale nie chce pokazać dobrze 2 skali Y na 1 wykresie
    df2 = df[df['Kontrakt'] == produkt]

    data = df2['Data']
    cena = df2['DKR']
    wolumen = df2['liczba transakcji']
    # Create figure with secondary y-axis
    fig = make_subplots(specs=[[{"secondary_y": True}]])
    # Add traces
    fig.add_trace(
        go.Scatter(x=data, y=cena, name="cena",mode='markers+lines', line=dict(color='blue')),
        secondary_y=False,
    )

    fig.add_trace(
        go.Bar(x=data, y=wolumen, name="Liczba transakcji",marker_color='gray', opacity=0.5),
        secondary_y=True,
    )
    #Dodaj interaktywny "krzyżak" do wykresu
    fig.update_traces(hoverinfo='text',
                      hovertemplate='<b>Data</b>: %{x}<br><b>Cena</b>: %{y:.2f}',
                      selector=dict(mode='markers+lines'))


    st.plotly_chart(fig)


st.set_page_config(layout="wide")   #rozciąga aplikacje na całą strone web
st.title("Wykres liniowy DKR, wraz z słupkami z wolumenem")

interactive= st.checkbox("włącz wykres interaktywny")

col1, col2, col3 = st.columns(3)
with col1:
    selected_Y=st.selectbox("Y - Produkty roczne",cale)
with col2:
    selected_Q = st.selectbox("Q - Produkty kwartalne", kwartaly)
with col3:
    selected_msc=st.selectbox("MSC - Produkty miesięczne", msc)

draw_y,draw_q,draw_msc= st.columns(3)
with draw_y:
    if interactive==True:
        draw_interactive((selected_Y))
    else:
        draw_chart(selected_Y)
with draw_q:
    if interactive==True:
        draw_interactive((selected_Q))
    else:
        draw_chart(selected_Q)
with draw_msc:
    if interactive==True:
        draw_interactive((selected_msc))
    else:
        draw_chart(selected_msc)

def pobierz_dane(url):
    response = requests.get(url)
    if response.status_code == 200:
        return response.text
    return None

def analizuj_dane(html):
    soup = BeautifulSoup(html, 'html.parser')
    tabelki = soup.find_all('table')
    dataframes = []
    for tabelka in tabelki:
        df = pd.read_html(str(tabelka), header=0, decimal=",", thousands='.')
        dataframes.extend(df)
    return dataframes
def aktualizacja():    #aktualizacja danych jako funkcja która wywołujemy przyciskiem
    wb = load_workbook(filename="abc.xlsx")
    ws = wb["a"]
    ostatni_wiersz = ws.max_row
    ostatnia_data = ws.cell(row=ostatni_wiersz,
                            column=1).value  # uchwycona ostatnia data, dla której są dane w pliku excel
    pl_holidays = holidays.Poland()

    # kod  na ostatni dzień roboczy
    dzisiaj = dt.date.today()
    delta1 = dt.timedelta(days=1)
    delta2 = dt.timedelta(days=2)
    ostatni_dzien = dzisiaj - delta1

    if ostatni_dzien.weekday() == 5:
        ostatni_dzien = ostatni_dzien - delta1
    elif ostatni_dzien.weekday() == 6:
        ostatni_dzien = ostatni_dzien - delta2

    ostatni_dzien_str = str(ostatni_dzien)

    for x in range(len(pl_holidays)):
        if ostatni_dzien_str in pl_holidays:
            ostatni_dzien = ostatni_dzien - delta1
            if ostatni_dzien.weekday() == 5:
                ostatni_dzien = ostatni_dzien - delta1
            elif ostatni_dzien.weekday() == 6:
                ostatni_dzien = ostatni_dzien - delta2
            ostatni_dzien_str = str(ostatni_dzien)

    weekdays = [5, 6]
    data_poczatkowa = dt.datetime.strptime(ostatnia_data,
                                           "%d-%m-%Y") + delta1  # trzeba do ostatniej daty dodać 1 dzień
    start_day = data_poczatkowa
    end_day = ostatni_dzien

    daterange = pd.date_range(start_day, end_day)
    for date in daterange:
        if date.weekday() not in weekdays and date.strftime("%Y-%m-%d") not in pl_holidays:
            dzien = date.strftime("%d-%m-%Y")
            url = 'https://tge.pl/energia-elektryczna-otf?dateShow=' + dzien + '&dateAction=prev'
            html=pobierz_dane(url)
            time.sleep(1)
            if html:
                df_list = analizuj_dane(html)
                for df in df_list:
                    # -----BASE & PEAK  ---
                    df.drop('Unnamed: 1', axis=1,inplace=True)  # Usunięcie kolumny 'Unnamed: 1'
                    df = df[~df['Kontrakt'].str.contains('OFFPEAK|H-PEAK|L-PEAK|W', case=False)] # Usunięcie rekordów, które w kolumnie 'Kontrakt' zawierają słowa 'offpeak', 'H-peak' lub 'L-Peak'
                    base = df #nazwa df jako base , ale kod zmieniony i zarówno base jak i peak przechowywany tutaj
                    base = base.iloc[:-1]  # Usunięcie ostatniego wiersza z tabeli (podsumowania)
                    base.insert(0, 'data', dzien)  # Dodanie daty w pierwszej kolumnie (0-zerowej)
                    if df['Kontrakt'].str.contains('BASE').any():
                        base['typ'] = "BASE"  # Dodajemy kolumnę Typ: "BASE" dla produktów z tabeli base, PEAK dla produktów z tabeli PEAK
                    else:
                        base['typ'] = "PEAK"
                    print(base)

                    wb = load_workbook(filename="abc.xlsx")
                    ws = wb["a"]
                    for x in dataframe_to_rows(base, index=False, header=False):
                        ws.append(x)  # Append dodaje dane do już istniejących w pliku
                    wb.save("abc.xlsx")


st.button('aktualizacja danych z TGE',on_click=aktualizacja)# Przycisk do aktualizacji danych, jeszcze do dopracowania wizualizacja np. popup z postepem, info o zakończeniu

#powtarzam część kodu żeby wyciagnac ostatnia datę poza funkcje i pokazac na strone
wb = load_workbook(filename="abc.xlsx")
ws = wb["a"]
ostatni_wiersz = ws.max_row
ostatnia_data = ws.cell(row=ostatni_wiersz,
                        column=1).value  # uchwycona ostatnia data, dla której są dane w pliku excel
st.write(f'dane aktualne na dzień: {ostatnia_data}')
#https://github.com/sliwka00/WykresyTGE-bs4/blob/master/wykresy.py        ścieżka do 'deploy' na streamlit ale wyskakuje błąd